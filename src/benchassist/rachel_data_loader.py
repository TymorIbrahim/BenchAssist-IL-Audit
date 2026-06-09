"""Load and normalise Rachel's pretrial-detention Excel datasets.

Reads three ``.xlsx`` files from ``rachel_data/``, normalises their
heterogeneous column schemas into a single flat structure, and exports
the combined 180-row corpus as CSV + JSONL ready for the audit pipeline.

Each Excel workbook contains four sheets:
    1. Synthetic dataset (60 rows, 6 variants × 10 base cases)
    2. Codebook
    3. Audit prompts
    4. Output coding template

Usage::

    from benchassist.rachel_data_loader import load_rachel_cases, export_rachel_cases
    cases = load_rachel_cases()          # list[CounterfactualCase]
    export_rachel_cases(cases)           # writes CSV + JSONL
"""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from benchassist.config import get_settings
from benchassist.schemas import CounterfactualCase

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
_DEFAULT_RACHEL_DIR = _PROJECT_ROOT / "rachel_data"

_EXCEL_FILES = [
    "synthetic_pretrial_detention_llm_audit_dataset.xlsx",
    "synthetic_pretrial_detention_llm_audit_dataset_2.xlsx",
    "synthetic_pretrial_detention_llm_audit_dataset_3.xlsx",
]

# Mapping from Rachel's variant_type labels → normalised variant_type
_VARIANT_TYPE_MAP: dict[str, str] = {
    # File 1 labels
    "V0_Control": "control",
    "V1_EthnicName": "ethnicity_proxy",
    "V2_Neighborhood": "neighborhood_proxy",
    "V3_Age": "age_proxy",
    "V4_Employment": "employment_proxy",
    "V5_FamilyStatus": "family_status_proxy",
    # File 2 & 3 labels (already normalised)
    "control": "control",
    "baseline": "control",
    "ethnicity_proxy": "ethnicity_proxy",
    "neighborhood_proxy": "neighborhood_proxy",
    "age_proxy": "age_proxy",
    "employment_proxy": "employment_proxy",
    "family_status_proxy": "family_status_proxy",
}

# Severity → urgency mapping
_SEVERITY_URGENCY_MAP: dict[str, str] = {
    "low": "low",
    "medium": "medium",
    "medium-low": "low",
    "medium-high": "high",
    "high": "high",
    "non-violent property/financial offense": "low",
}


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _read_excel_sheet(path: Path, sheet_index: int = 0) -> list[dict[str, Any]]:
    """Read an Excel sheet and return rows as list of dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[sheet_index]]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []
        headers = [str(h or "").strip() for h in rows[0]]
        return [
            {headers[i]: cell for i, cell in enumerate(row) if i < len(headers)}
            for row in rows[1:]
        ]
    finally:
        wb.close()


def _read_prompts_sheet(path: Path) -> list[dict[str, str]]:
    """Read the Prompts / Audit Prompts sheet from an Excel file."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        prompt_sheet = None
        for name in wb.sheetnames:
            if "prompt" in name.lower():
                prompt_sheet = name
                break
        if prompt_sheet is None:
            return []
        ws = wb[prompt_sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []
        headers = [str(h or "").strip() for h in rows[0]]
        return [
            {headers[i]: str(cell or "").strip() for i, cell in enumerate(row) if i < len(headers)}
            for row in rows[1:]
            if any(cell for cell in row)
        ]
    finally:
        wb.close()


def _normalise_severity_to_urgency(severity: Any) -> str:
    """Map offense severity to an urgency level."""
    if severity is None:
        return "medium"
    text = str(severity).strip().lower()
    return _SEVERITY_URGENCY_MAP.get(text, "medium")


def _build_prompt_from_fields(row: dict[str, Any]) -> str:
    """Assemble a structured English prompt from individual case fields.

    Used for File 3 rows (and as fallback for File 2 rows that lack
    ``prompt_ready_case``).
    """
    parts = []

    def _add(label: str, key: str) -> None:
        val = row.get(key)
        if val is not None and str(val).strip():
            parts.append(f"{label}: {str(val).strip()}")

    _add("Suspect name", "suspect_name")
    _add("Age", "age")
    _add("Gender", "gender")
    _add("Neighborhood", "neighborhood")
    _add("Employment status", "employment_status")
    _add("Family status", "family_status")
    _add("Offense type", "offense_type")
    _add("Offense severity", "offense_severity")

    # Case summary / facts
    summary = row.get("case_summary")
    if summary and str(summary).strip():
        parts.append(f"Case summary: {str(summary).strip()}")

    _add("Police claim", "police_claim")
    _add("Police framing", "police_framing")
    _add("Defense claim", "defense_claim")
    _add("Defense argument", "defense_argument")
    _add("Prosecution request", "prosecution_request")
    _add("Prior record", "prior_record")
    _add("Evidence strength", "evidence_strength")
    _add("Investigation needs", "investigation_needs")
    _add("Proposed alternative", "proposed_alternative")
    _add("Alternative detention available", "alternative_detention_available")

    return "\n".join(parts)


def _normalise_variant_type(raw: Any) -> str:
    """Normalise variant_type / variation_type to a canonical label."""
    if raw is None:
        return "control"
    text = str(raw).strip()
    return _VARIANT_TYPE_MAP.get(text, text)


def _normalise_demographic_cue(row: dict[str, Any]) -> str:
    """Extract a demographic-cue description from varied column names."""
    # Try File 1's columns
    for key in ("manipulated_variable", "changed_variable"):
        val = row.get(key)
        if val and str(val).strip().lower() not in ("none", ""):
            return str(val).strip()
    # Try inferred group marker / ethnic signal / audit_note
    for key in ("inferred_group_marker", "ethnic_signal", "audit_note",
                "expected_bias_test", "expected_audit_focus"):
        val = row.get(key)
        if val and str(val).strip():
            return str(val).strip()
    return "none"


def _row_to_counterfactual_case(
    row: dict[str, Any],
    *,
    file_index: int,
) -> CounterfactualCase:
    """Convert a normalised Excel row into a ``CounterfactualCase``."""
    # Case ID
    case_id = str(
        row.get("case_id") or row.get("record_id") or ""
    ).strip()
    base_case_id = str(row.get("base_case_id", "")).strip()
    variant_id = case_id  # unique row ID already

    # Variant type
    variant_type = _normalise_variant_type(
        row.get("variant_type") or row.get("variation_type")
    )

    # Demographic cue
    demographic_cue = _normalise_demographic_cue(row)

    # Input text
    if file_index == 0 and row.get("prompt_input_hebrew"):
        input_text = str(row["prompt_input_hebrew"]).strip()
        language = "he"
    elif file_index == 1 and row.get("prompt_ready_case"):
        input_text = str(row["prompt_ready_case"]).strip()
        language = "en"
    else:
        input_text = _build_prompt_from_fields(row)
        language = "en"

    # Urgency from severity
    expected_urgency = _normalise_severity_to_urgency(
        row.get("offense_severity")
    )

    return CounterfactualCase(
        case_id=base_case_id or case_id,
        variant_id=variant_id,
        variant_type=variant_type,
        demographic_cue=demographic_cue,
        language=language,
        transformation_style="",
        input_text=input_text,
        expected_urgency=expected_urgency,
        expected_direction="",  # not applicable to detention
        strict_counterfactual_candidate=(variant_type != "control"),
        framing_axis="",
        framing_direction="",
        dataset_mode="synthetic_controlled",
        source_type="rachel_excel",
        source_dataset=_EXCEL_FILES[file_index],
        is_synthetic=True,
        legal_area="pretrial_detention",
    )


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def load_rachel_cases(
    rachel_dir: Path | str | None = None,
) -> list[CounterfactualCase]:
    """Load and normalise all 180 cases from Rachel's Excel files.

    Args:
        rachel_dir: Directory containing the 3 ``.xlsx`` files.
            Defaults to ``<project_root>/rachel_data/``.

    Returns:
        A flat list of 180 :class:`CounterfactualCase` instances.
    """
    directory = Path(rachel_dir) if rachel_dir else _DEFAULT_RACHEL_DIR
    all_cases: list[CounterfactualCase] = []

    for file_index, filename in enumerate(_EXCEL_FILES):
        fpath = directory / filename
        if not fpath.exists():
            logger.warning("Rachel data file not found: %s", fpath)
            continue

        rows = _read_excel_sheet(fpath, sheet_index=0)
        logger.info(
            "Loaded %d rows from %s (sheet: %s)",
            len(rows),
            filename,
            "dataset",
        )

        for row in rows:
            try:
                case = _row_to_counterfactual_case(row, file_index=file_index)
                all_cases.append(case)
            except Exception:
                logger.exception(
                    "Failed to convert row from %s: %s",
                    filename,
                    row.get("case_id") or row.get("record_id"),
                )

    logger.info(
        "Total Rachel cases loaded: %d (expected 180)", len(all_cases)
    )
    return all_cases


def load_rachel_prompts(
    rachel_dir: Path | str | None = None,
) -> list[dict[str, str]]:
    """Load audit prompts from all three Excel files.

    Returns:
        A deduplicated list of prompt dicts with 'name' and 'text' keys.
    """
    directory = Path(rachel_dir) if rachel_dir else _DEFAULT_RACHEL_DIR
    all_prompts: list[dict[str, str]] = []
    seen_names: set[str] = set()

    for filename in _EXCEL_FILES:
        fpath = directory / filename
        if not fpath.exists():
            continue
        prompts = _read_prompts_sheet(fpath)
        for p in prompts:
            name = (
                p.get("Prompt Name")
                or p.get("Prompt name")
                or p.get("name", "")
            ).strip()
            text = (
                p.get("Prompt Text")
                or p.get("Prompt text")
                or p.get("text", "")
            ).strip()
            if name and name not in seen_names:
                seen_names.add(name)
                all_prompts.append({"name": name, "text": text})

    return all_prompts


def export_rachel_cases(
    cases: list[CounterfactualCase] | None = None,
    *,
    output_dir: Path | str | None = None,
    rachel_dir: Path | str | None = None,
    basename: str = "rachel_detention_cases",
) -> tuple[Path, Path]:
    """Export Rachel cases to CSV and JSONL.

    Args:
        cases: Pre-loaded cases (will load if ``None``).
        output_dir: Target directory (defaults to ``data/audit/``).
        rachel_dir: Source directory for Excel files.
        basename: Output file basename (without extension).

    Returns:
        Tuple of (csv_path, jsonl_path).
    """
    if cases is None:
        cases = load_rachel_cases(rachel_dir)

    settings = get_settings()
    out_dir = Path(output_dir) if output_dir else (settings.DATA_DIR / "audit")
    out_dir.mkdir(parents=True, exist_ok=True)

    csv_path = out_dir / f"{basename}.csv"
    jsonl_path = out_dir / f"{basename}.jsonl"

    # CSV
    df = pd.DataFrame([c.model_dump() for c in cases])
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    logger.info("Wrote %d cases to %s", len(cases), csv_path)

    # JSONL
    with open(jsonl_path, "w", encoding="utf-8") as fh:
        for case in cases:
            fh.write(case.model_dump_json() + "\n")
    logger.info("Wrote %d cases to %s", len(cases), jsonl_path)

    return csv_path, jsonl_path


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------


def main() -> None:
    """CLI: load Rachel data and export to data/audit/."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Load Rachel's pretrial detention Excel data and export as CSV/JSONL."
    )
    parser.add_argument(
        "--rachel-dir",
        type=Path,
        default=None,
        help="Directory containing the 3 .xlsx files (default: rachel_data/).",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="Output directory for CSV/JSONL (default: data/audit/).",
    )
    parser.add_argument(
        "--basename",
        default="rachel_detention_cases",
        help="Output file basename (default: rachel_detention_cases).",
    )
    args = parser.parse_args()

    logging.basicConfig(level="INFO", format="%(levelname)s: %(message)s")

    cases = load_rachel_cases(args.rachel_dir)
    print(f"Loaded {len(cases)} cases")

    # Summary
    base_ids = {c.case_id for c in cases}
    print(f"  Unique base case IDs: {len(base_ids)}")
    variant_types = {c.variant_type for c in cases}
    print(f"  Variant types: {sorted(variant_types)}")
    languages = {c.language for c in cases}
    print(f"  Languages: {sorted(languages)}")

    csv_path, jsonl_path = export_rachel_cases(
        cases,
        output_dir=args.output_dir,
        basename=args.basename,
    )
    print(f"\nExported:")
    print(f"  → {csv_path}")
    print(f"  → {jsonl_path}")

    # Also export prompts
    prompts = load_rachel_prompts(args.rachel_dir)
    if prompts:
        print(f"\nAudit prompts found ({len(prompts)}):")
        for p in prompts:
            print(f"  • {p['name']}: {p['text'][:80]}...")


if __name__ == "__main__":
    main()
